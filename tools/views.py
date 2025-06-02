import json
import re

from difflib import get_close_matches
from django.shortcuts import render, get_object_or_404, redirect
from django.forms import formset_factory
from django.http import HttpResponse, JsonResponse
from .models import Material, Project, BuildingType, Document
from .forms import *
from google import genai
from django.contrib import messages
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import io

MaterialFormSet = formset_factory(MaterialQuantityForm, extra=1)
client = genai.Client(api_key='AIzaSyCbHyl224pNyX_HOnYFtjwQr_o2nH8GgLU')


def generate_carbon_footprint_insight(breakdown, total_emission):
    prompt = (
        "You are an AI sustainability assistant analyzing carbon emissions from construction materials. "
        "You are given a list of materials, each with a name, quantity used (in kg), and total carbon emissions (in kg CO2). "
        "For each material, provide detailed insights about its environmental impact and suggest sustainable alternatives. "
        "Also include a total emission summary and a high-level insight for the overall material usage."

        "\n\nRespond ONLY in the following JSON format and include nothing else:\n"
        """{
          "materials": [
            {
              "material": "string",
              "quantity": float,
              "emission": float,
              "emission_per_unit": float,
              "environmental_impact": "string",
              "alternatives": [
                {
                  "material": "string",
                  "description": "string",
                  "emission_factor": "string or float",
                  "considerations": "string"
                }
              ],
              "recommendations": "string"
            }
          ],
          "total_emission": float,
          "summary_insight": "string"
        }"""

        "\n\nHere is the input data of materials with their quantities and emissions:\n"
        f"{breakdown}, with total emission: {total_emission}.\n"
        "Only return the valid JSON response as described above."
    )
    response = client.models.generate_content(
    model='gemini-2.0-flash-001', contents=prompt
    )

    raw_text = response.text.strip()

    # Sometimes the model includes triple quotes or markdown-style formatting; remove them
    if raw_text.startswith("```json"):
        raw_text = raw_text.lstrip("```json").rstrip("```").strip()

    try:
        carbon_data = json.loads(raw_text)
    except json.JSONDecodeError as e:
        print("Failed to parse JSON:", e)
        waste_data = {}
    print("waste data", carbon_data)
    return carbon_data


def generate_waste_forecasting(building_size, material, waste_factor,
                construction_method, project_phase, climate):
    prompt = (
                "Based on the following construction context, provide smart waste reduction tips and a brief forecast:"
                "Provide output in json format and do not add any other comments."
                f"- Building size: {building_size} sq ft"
                f"- Materials used are: {material}"
                f"- Waste factor: {waste_factor}%"
                f"- Construction method: {construction_method}"
                f"- Project phase: {project_phase}"
                f"- Climate: {climate}"
                
                """Respond only in JSON with:
                {
                  "predicted_waste_kg": ...,
                  "waste_reduction_tips": [
                    "...",
                    "..."
                  ],
                  "expected_material_wastage_kg": float,
                  "reason_for_wastage": str
                }
                """)
    response = client.models.generate_content(
    model='gemini-2.0-flash-001', contents=prompt
    )

    raw_text = response.text.strip()

    # Sometimes the model includes triple quotes or markdown-style formatting; remove them
    if raw_text.startswith("```json"):
        raw_text = raw_text.lstrip("```json").rstrip("```").strip()

    try:
        waste_data = json.loads(raw_text)
    except json.JSONDecodeError as e:
        print("Failed to parse JSON:", e)
        waste_data = {}
    print("waste data", waste_data)
    return waste_data


def generate_design_suggestion_gemini(sunlight, airflow, energy_budget, size):
    prompt = (

        """You are an AI assistant for sustainable architecture.
            
            Given the following building design constraints:
            - Sunlight in hrs/day
            - Airflow in cfm
            - Energy Budget in kWh/year
            - Building Size in square ft
            
            Generate optimized green building suggestions, formatted as structured JSON.
            
            Response Format:
            {
              "passive_design_strategies": ["..."],
              "recommended_materials": [
                {"name": "...", "reason": "..."},
                ...
              ],
              "energy_efficiency_tips": ["..."],
              "additional_notes": "..."
            }
            
            Make sure all items are concise, relevant, and environmentally responsible.
            """ +
            f"Suggest sustainable building design strategies for:\n"
            f"- Sunlight: {sunlight} hours/day\n"
            f"- Airflow: {airflow} cfm\n"
            f"- Energy Budget: {energy_budget} kWh/year\n"
            f"- Size: {size} square meters\n"
            f"Include passive design ideas, material suggestions, and energy efficiency tips."
            f"Respond with only a valid json object without any additional commentary."
    )

    response = client.models.generate_content(
    model='gemini-2.0-flash-001', contents=prompt
    )

    raw_text = response.text.strip()

    # Sometimes the model includes triple quotes or markdown-style formatting; remove them
    if raw_text.startswith("```json"):
        raw_text = raw_text.lstrip("```json").rstrip("```").strip()

    try:
        design_data = json.loads(raw_text)
    except json.JSONDecodeError as e:
        print("Failed to parse JSON:", e)
        design_data = {}
    return design_data


def analyze_pdf_from_file(file_path, prompt):
    """
    Uploads a PDF file and asks Gemini to answer a question based on its content.
    """
    # Upload the PDF file
    uploaded_file = client.files.upload(file=file_path)

    response = client.models.generate_content(
        model='gemini-2.0-flash-001',
        contents=[prompt, uploaded_file]
    )
    return response

def dashboard(request):
    projects = Project.objects.all().order_by('-created_at')
    return render(request, "dashboard.html", {"projects": projects})


def generate_leed_certification_insight(project):
    # Gather existing insights from project
    carbon = project.carbon_insight or {}
    waste = project.waste_insight or {}
    design = project.design_insight or {}

    # Convert JSON strings if needed
    import json
    for insight in [carbon, waste, design]:
        if isinstance(insight, str):
            try:
                insight = json.loads(insight)
            except:
                insight = {}

    # Compose the prompt with all insights embedded
    prompt = f"""
    You are an AI assistant helping companies find their LEED v5 certification score for LEED v5 certificate tasked with generating a combined insight and forecasting for a construction project.

    You are required to provide proper LEED v5 certification analysis for the following project:{project}
    
    Please analyze all data and provide a combined JSON output with:
    
    - A human-readable combined insight summary containing two parts
        1. LEED score for choosen LEED certification by the user and current data available will be value for key "score" and reason for its calculation will be value of key "reasons", reasons value will be list of strings.
        2. Suggestions for improving all factors to maximise LEED score and then predicted LEED score once user follows all suggestion properly. Here also provide result in two keys 'score' and 'reasons'. Similar to above point.
    - In forecasting make sure no past year is provide , forecasting will be made from current year to current year + 1 year
    - In the final insight also add the percentage reduction or increase in cost if adviced material are
     used.
    - Forecasting information for future emissions, waste, and energy efficiency.
    - Graph metrics data suitable for bar graphs, pie charts, or line charts, including:
      - total_carbon_emission (float)
      - predicted_waste_kg (float)
      - energy_efficiency_score (float, 0-100)
      - leed_score (int, 0-110) (USGBC LEED Certificate score according to your analysis)
      - leed_recommendation_grade (str)
    Note: leed_recommendation_grade is string providing grades for particular range of score 110 Points [Failed(0-39 Points),Certified(40-49 Points),Silver(50-59 Points), Gold(60-79Points),Platinum(80+ Points)],
    - To justify the current score you ahve provided also provide the scorecard details in key 'scorecard' so that user can check in which categories they have received or not received any score.
    - Scorecard has been provided to you in json format with max points for each category, whatever current score you provide you also need to modify the scorecard so that the score you have provided is justified. All the keys mentioned in that scorecard should be available in the scorecard json provided by you.
    Return ONLY the JSON response exactly in this format, with no extra text:
    
    {{
      "combined_insight": {{
        "current_scenario": {{'score': int, 'reasons': list of strings}},
        "suggestion": {{'score': int, 'reasons': list of strings}}
      }},
      "forecasting": {{
        "carbon_emission_trend": [{{"year": int, "emission": float}}, ...],
        "waste_trend": [{{"year": int, "waste": float}}, ...],
        "energy_efficiency_trend": [{{"year": int, "score": float}}, ...]
      }},
      "graph_metrics": {{
        "total_carbon_emission": float,
        "predicted_waste_kg": float,
        "energy_efficiency_score": float,
        "sustainability_recommendation_score": int,
        "sustainability_recommendation_grade": string
      }},
      "scorecard": {{
        "Integrative Process, Planning and Assessments (IP): {{ 
            "integrative_design_process": int,
            "green_leases": int 
        }},
        "Location and Transportation": {{
            "Sensitive Land Protection": int,
            ...and so on
        }},... and so on
      }}
    }}
    """

    file_path = "LEED Scorecards and Reference Guides"
    if project.leed_certification == "LEED BD+C(New Construction)":
        scorecard_json = {
                              "Integrative Process, Planning and Assessments (IP)": {
                                "total_points": 1,
                                "criteria": {
                                  "IPp1": { "name": "Climate Resilience Assessment", "points": "Required" },
                                  "IPp2": { "name": "Human Impact Assessment", "points": "Required" },
                                  "IPp3": { "name": "Carbon Assessment", "points": "Required" },
                                  "IPc1": { "name": "Integrative Design Process", "points": 1 }
                                }
                              },
                              "Location and Transportation (LT)": {
                                "total_points": 15,
                                "criteria": {
                                  "LTc1": { "name": "Sensitive Land Protection", "points": 1 },
                                  "LTc2": { "name": "Equitable Development", "points": 2 },
                                  "LTc3": { "name": "Compact and Connected Development", "points": 6 },
                                  "LTc4": { "name": "Transportation Demand Management", "points": 4 },
                                  "LTc5": { "name": "Electric Vehicles", "points": 2 }
                                }
                              },
                              "Sustainable Sites (SS)": {
                                "total_points": 11,
                                "criteria": {
                                  "SSp1": { "name": "Minimize Site Disturbance", "points": "Required" },
                                  "SSc1": { "name": "Biodiverse Habitat", "points": 2 },
                                  "SSc2": { "name": "Accessible Outdoor Space", "points": 1 },
                                  "SSc3": { "name": "Rainwater Management", "points": 3 },
                                  "SSc4": { "name": "Enhanced Resilient Site Design", "points": 2 },
                                  "SSc5": { "name": "Heat Island Reduction", "points": 2 },
                                  "SSc6": { "name": "Light Pollution Reduction", "points": 1 }
                                }
                              },
                              "Water Efficiency (WE)": {
                                "total_points": 9,
                                "criteria": {
                                  "WEp1": { "name": "Water Metering and Reporting", "points": "Required" },
                                  "WEp2": { "name": "Minimum Water Efficiency", "points": "Required" },
                                  "WEc1": { "name": "Water Metering and Leak Detection", "points": 1 },
                                  "WEc2": { "name": "Enhanced Water Efficiency", "points": 8 }
                                }
                              },
                              "Energy and Atmosphere (EA)": {
                                "total_points": 33,
                                "criteria": {
                                  "EAp1": { "name": "Operational Carbon Projection and Decarbonization Plan", "points": "Required" },
                                  "EAp2": { "name": "Minimum Energy Efficiency", "points": "Required" },
                                  "EAp3": { "name": "Fundamental Commissioning", "points": "Required" },
                                  "EAp4": { "name": "Energy Metering and Reporting", "points": "Required" },
                                  "EAp5": { "name": "Fundamental Refrigerant Management", "points": "Required" },
                                  "EAc1": { "name": "Electrification", "points": 5 },
                                  "EAc2": { "name": "Reduce Peak Thermal Loads", "points": 5 },
                                  "EAc3": { "name": "Enhanced Energy Efficiency", "points": 10 },
                                  "EAc4": { "name": "Renewable Energy", "points": 5 },
                                  "EAc5": { "name": "Enhanced Commissioning", "points": 4 },
                                  "EAc6": { "name": "Grid Interactive", "points": 2 },
                                  "EAc7": { "name": "Enhanced Refrigerant Management", "points": 2 }
                                }
                              },
                              "Materials and Resources (MR)": {
                                "total_points": 18,
                                "criteria": {
                                  "MRp1": { "name": "Planning for Zero Waste Operations", "points": "Required" },
                                  "MRp2": { "name": "Quantify and Assess Embodied Carbon", "points": "Required" },
                                  "MRc1": { "name": "Building and Materials Reuse", "points": 3 },
                                  "MRc2": { "name": "Reduce Embodied Carbon", "points": 6 },
                                  "MRc3": { "name": "Low-Emitting Materials", "points": 2 },
                                  "MRc4": { "name": "Building Product Selection and Procurement", "points": 5 },
                                  "MRc5": { "name": "Construction and Demolition Waste Diversion", "points": 2 }
                                }
                              },
                              "Indoor Environmental Quality (EQ)": {
                                "total_points": 13,
                                "criteria": {
                                  "EQp1": { "name": "Construction Management", "points": "Required" },
                                  "EQp2": { "name": "Fundamental Air Quality", "points": "Required" },
                                  "EQp3": { "name": "No Smoking or Vehicle Idling", "points": "Required" },
                                  "EQc1": { "name": "Enhanced Air Quality", "points": 1 },
                                  "EQc2": { "name": "Occupant Experience", "points": 7 },
                                  "EQc3": { "name": "Accessibility and Inclusion", "points": 1 },
                                  "EQc4": { "name": "Resilient Spaces", "points": 2 },
                                  "EQc5": { "name": "Air Quality Testing and Monitoring", "points": 2 }
                                }
                              },
                              "Project Priorities (PR)": {
                                "total_points": 10,
                                "criteria": {
                                  "PRc1": { "name": "Project Priorities", "points": 9 },
                                  "PRc2": { "name": "LEED AP", "points": 1 }
                                }
                              },
                              "total_points": 110
                        }

        file_path += "/LEED v5 BD+C Reference Guide_Launch Edition.pdf"
    elif project.leed_certification == "LEED BD+C(Core and Shell)":
        scorecard_json = {
            "Integrative Process, Planning and Assessments (IP)": {
                "total_points": 7,
                "criteria": {
                    "IPp1": {"name": "Climate Resilience Assessment", "points": "Required"},
                    "IPp2": {"name": "Human Impact Assessment", "points": "Required"},
                    "IPp3": {"name": "Carbon Assessment", "points": "Required"},
                    "IPp4": {"name": "Tenant Guidelines", "points": "Required"},
                    "IPc1": {"name": "Integrative Design Process", "points": 1},
                    "IPc2": {"name": "Green Leases", "points": 6}
                }
            },
            "Location and Transportation (LT)": {
                "total_points": 15,
                "criteria": {
                    "LTc1": {"name": "Sensitive Land Protection", "points": 1},
                    "LTc2": {"name": "Equitable Development", "points": 2},
                    "LTc3": {"name": "Compact and Connected Development", "points": 6},
                    "LTc4": {"name": "Transportation Demand Management", "points": 4},
                    "LTc5": {"name": "Electric Vehicles", "points": 2}
                }
            },
            "Sustainable Sites (SS)": {
                "total_points": 11,
                "criteria": {
                    "SSp1": {"name": "Minimize Site Disturbance", "points": "Required"},
                    "SSc1": {"name": "Biodiverse Habitat", "points": 2},
                    "SSc2": {"name": "Accessible Outdoor Space", "points": 1},
                    "SSc3": {"name": "Rainwater Management", "points": 3},
                    "SSc4": {"name": "Enhanced Resilient Site Design", "points": 2},
                    "SSc5": {"name": "Heat Island Reduction", "points": 2},
                    "SSc6": {"name": "Light Pollution Reduction", "points": 1}
                }
            },
            "Water Efficiency (WE)": {
                "total_points": 8,
                "criteria": {
                    "WEp1": {"name": "Water Metering and Reporting", "points": "Required"},
                    "WEp2": {"name": "Minimum Water Efficiency", "points": "Required"},
                    "WEc1": {"name": "Water Metering and Leak Detection", "points": 1},
                    "WEc2": {"name": "Enhanced Water Efficiency", "points": 7}
                }
            },
            "Energy and Atmosphere (EA)": {
                "total_points": 27,
                "criteria": {
                    "EAp1": {"name": "Operational Carbon Projection and Decarbonization Plan", "points": "Required"},
                    "EAp2": {"name": "Minimum Energy Efficiency", "points": "Required"},
                    "EAp3": {"name": "Fundamental Commissioning", "points": "Required"},
                    "EAp4": {"name": "Energy Metering and Reporting", "points": "Required"},
                    "EAp5": {"name": "Fundamental Refrigerant Management", "points": "Required"},
                    "EAc1": {"name": "Electrification", "points": 4},
                    "EAc2": {"name": "Reduce Peak Thermal Loads", "points": 5},
                    "EAc3": {"name": "Enhanced Energy Efficiency", "points": 7},
                    "EAc4": {"name": "Renewable Energy", "points": 4},
                    "EAc5": {"name": "Enhanced Commissioning", "points": 3},
                    "EAc6": {"name": "Grid Interactive", "points": 2},
                    "EAc7": {"name": "Enhanced Refrigerant Management", "points": 2}
                }
            },
            "Materials and Resources (MR)": {
                "total_points": 21,
                "criteria": {
                    "MRp1": {"name": "Planning for Zero Waste Operations", "points": "Required"},
                    "MRp2": {"name": "Quantify and Assess Embodied Carbon", "points": "Required"},
                    "MRc1": {"name": "Building and Materials Reuse", "points": 5},
                    "MRc2": {"name": "Reduce Embodied Carbon", "points": 8},
                    "MRc3": {"name": "Low-Emitting Materials", "points": 1},
                    "MRc4": {"name": "Building Product Selection and Procurement", "points": 5},
                    "MRc5": {"name": "Construction and Demolition Waste Diversion", "points": 2}
                }
            },
            "Indoor Environmental Quality (EQ)": {
                "total_points": 11,
                "criteria": {
                    "EQp1": {"name": "Construction Management", "points": "Required"},
                    "EQp2": {"name": "Fundamental Air Quality", "points": "Required"},
                    "EQp3": {"name": "No Smoking or Vehicle Idling", "points": "Required"},
                    "EQc1": {"name": "Enhanced Air Quality", "points": 1},
                    "EQc2": {"name": "Occupant Experience", "points": 7},
                    "EQc3": {"name": "Accessibility and Inclusion", "points": 1},
                    "EQc4": {"name": "Resilient Spaces", "points": 2},
                    "EQc5": {"name": "Air Quality Testing and Monitoring", "points": 0}
                }
            },
            "Project Priorities (PR)": {
                "total_points": 10,
                "criteria": {
                    "PRc1": {"name": "Project Priorities", "points": 9},
                    "PRc2": {"name": "LEED AP", "points": 1}
                }
            },
            "total_points": 110
        }
        file_path += "/LEED v5 BD+C Reference Guide_Launch Edition.pdf"
    elif project.leed_certification == "LEED ID+C":
        scorecard_json = {
                              "Integrative Process, Planning, and Assessments (IP)": {
                                "total_points": 1,
                                "criteria": {
                                  "IPp1 Climate Resilience Assessment": "Required",
                                  "IPp2 Human Impact Assessment": "Required",
                                  "IPp3 Carbon Assessment": "Required",
                                  "IPc1 Integrative Design Process": 1
                                }
                              },
                              "Location and Transportation (LT)": {
                                "total_points": 14,
                                "criteria": {
                                  "LTc1 Compact and Connected Development": 8,
                                  "LTc2 Transportation Demand Management": 4,
                                  "LTc3 Electric Vehicles": 2
                                }
                              },
                              "Water Efficiency (WE)": {
                                "total_points": 10,
                                "criteria": {
                                  "WEp1 Minimum Water Efficiency": "Required",
                                  "WEc1 Water Metering and Leak Detection": 2,
                                  "WEc2 Enhanced Water Efficiency": 8
                                }
                              },
                              "Energy and Atmosphere (EA)": {
                                "total_points": 31,
                                "criteria": {
                                  "EAp1 Estimated Energy Use and Operational Carbon Projection": "Required",
                                  "EAp2 Minimum Energy Efficiency": "Required",
                                  "EAp3 Fundamental Commissioning": "Required",
                                  "EAp4 Energy Metering and Reporting": "Required",
                                  "EAp5 Fundamental Refrigerant Management": "Required",
                                  "EAc1 Electrification": 5,
                                  "EAc2 Enhanced Energy Efficiency": 12,
                                  "EAc3 Renewable Energy": 5,
                                  "EAc4 Enhanced Commissioning": 4,
                                  "EAc5 Grid Interactive": 3,
                                  "EAc6 Enhanced Refrigerant Management": 2
                                }
                              },
                              "Materials and Resources (MR)": {
                                "total_points": 26,
                                "criteria": {
                                  "MRp1 Planning for Zero Waste Operations": "Required",
                                  "MRp2 Quantify and Assess Embodied Carbon": "Required",
                                  "MRc1 Interior Materials Reuse": 4,
                                  "MRc2 Reduce Embodied Carbon": 4,
                                  "MRc3 Low-Emitting Materials": 4,
                                  "MRc4 Building Product Selection and Procurement": 10,
                                  "MRc5 Construction and Demolition Waste Diversion": 4
                                }
                              },
                              "Indoor Environmental Quality (EQ)": {
                                "total_points": 18,
                                "criteria": {
                                  "EQp1 Construction Management": "Required",
                                  "EQp2 Fundamental Air Quality": "Required",
                                  "EQp3 No Smoking": "Required",
                                  "EQc1 Enhanced Air Quality": 2,
                                  "EQc2 Occupant Experience": 7,
                                  "EQc3 Accessibility and Inclusion": 2,
                                  "EQc4 Resilient Spaces": 3,
                                  "EQc5 Air Quality Testing and Monitoring": 4
                                }
                              },
                              "Project Priorities (PR)": {
                                "total_points": 10,
                                "criteria": {
                                  "PRc1 Project Priorities": 9,
                                  "PRc2 LEED AP": 1
                                }
                              },
                              "Total Possible Points": 110
                        }
        file_path += "/LEED v5 ID+C Reference Guide_Launch Edition 1.pdf"
    elif project.leed_certification == "LEED O+M":
        scorecard_json = {
                              "Integrative Process, Planning and Assessments (IP)": {
                                "total_points": 2,
                                "criteria": {
                                  "IPp1 Climate Resilience Assessment": "Required",
                                  "IPp2 Human Impact Assessment": "Required",
                                  "IPp3 Operations Assessment and Policy": "Required",
                                  "IPp4 Current Facilities Requirements and O+M Plan": "Required",
                                  "IPc1 Operational Planning for Resilience": 1,
                                  "IPc2 Worker Safety and Training": 1
                                }
                              },
                              "Location and Transportation (LT)": {
                                "total_points": 8,
                                "criteria": {
                                  "LTc1 Sustainable Transportation Performance": 6,
                                  "LTc2 Transportation Demand Management": 1,
                                  "LTc3 Electric Vehicles": 1
                                }
                              },
                              "Sustainable Sites (SS)": {
                                "total_points": 2,
                                "criteria": {
                                  "SSc1 Heat Island Reduction": 1,
                                  "SSc2 Light Pollution and Bird Collision Reduction": 1
                                }
                              },
                              "Water Efficiency (WE)": {
                                "total_points": 15,
                                "criteria": {
                                  "WEp1 Water Metering and Reporting": "Required",
                                  "WEc1 Water Efficiency Performance": 14,
                                  "WEc2 Advanced Water Metering": 1
                                }
                              },
                              "Energy and Atmosphere (EA)": {
                                "total_points": 34,
                                "criteria": {
                                  "EAp1 Carbon Projection from Energy Use": "Required",
                                  "EAp2 Energy Monitoring and Reporting": "Required",
                                  "EAp3 Minimum Energy Performance": "Required",
                                  "EAp4 Fundamental Refrigerant Management": "Required",
                                  "EAc1 Greenhouse Gas Emissions Reduction Performance": 12,
                                  "EAc2 Optimized Energy Performance": 12,
                                  "EAc3 Enhanced Refrigerant Management Performance": 2,
                                  "EAc4 Peak Load Reduction Performance": 1,
                                  "EAc5 Decarbonization and Efficiency Plans": 4,
                                  "EAc6 Peak Load Management": 1,
                                  "EAc7 Commissioning": 2
                                }
                              },
                              "Materials and Resources (MR)": {
                                "total_points": 13,
                                "criteria": {
                                  "MRc1 Waste Reduction Performance": 12,
                                  "MRc2 Waste Reduction Strategies": 1
                                }
                              },
                              "Indoor Environmental Quality (EQ)": {
                                "total_points": 26,
                                "criteria": {
                                  "EQp1 Verification of Ventilation and Filtration": "Required",
                                  "EQp2 No Smoking": "Required",
                                  "EQc1 Indoor Air Quality Performance": 10,
                                  "EQc2 Ventilation Performance": 5,
                                  "EQc3 Occupant Experience Performance": 3,
                                  "EQc4 Facility Stewardship Performance": 3,
                                  "EQc5 Air Filtration": 1,
                                  "EQc6 Resilient Spaces": 1,
                                  "EQc7 Green Cleaning": 2,
                                  "EQc8 Integrated Pest Management": 1
                                }
                              },
                              "Project Priorities (PR)": {
                                "total_points": 10,
                                "criteria": {
                                  "PRc1 Project Priorities": 10
                                }
                              },
                              "Total Possible Points": 110
                        }
        file_path += "/LEED v5 O+M Reference Guide_Launch Edition 1.pdf"
    else:
        return ""

    prompt += f"This is the scorecard for this category, make sure the keys you provide for leed scorecard should be exactly matching with this json just modify the values according to actual values based on current data we have {scorecard_json}"
    response = analyze_pdf_from_file(
        file_path=file_path,
        prompt=prompt
    )

    raw_text = response.text.strip()

    # Clean possible markdown or code fences
    if raw_text.startswith("```json"):
        raw_text = raw_text.lstrip("```json").rstrip("```").strip()

    try:
        combined_data = json.loads(raw_text)
    except json.JSONDecodeError as e:
        print("JSON parse error:", e)
        combined_data = {}
    print("insight", combined_data.get("combined_insight"))
    # Save back to project
    project.leed_certification_insight = combined_data.get("combined_insight", "")
    project.leed_combined_forecasting = combined_data.get("forecasting", {})
    project.leed_graph_metrics = combined_data.get("graph_metrics", {})
    project.leed_scorecard = combined_data.get("scorecard", {})
    project.save()

    return combined_data

def is_number(value):
    return isinstance(value, (int, float))


def download_leed_scorecard(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    leed_scorecard = project.leed_scorecard  # Replace this with your JSON source

    wb = Workbook()
    ws = wb.active
    ws.title = "LEED Scorecard"

    # Header
    ws.append(["Category", "Total Points", "Criterion Code", "Criterion Name", "Points"])

    for category, content in leed_scorecard.items():
        if is_number(content):
            continue
        total_points = content.get("total_points", None)
        criteria = content.get("criteria", {})

        for full_key, point_value in criteria.items():
            ws.append([category, total_points, full_key, point_value.get("name"),
                       point_value.get("points")])

    # Save Excel to memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    response = HttpResponse(
        file_stream,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=leed_scorecard.xlsx'
    return response


def download_well_scorecard(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    well_scorecard = project.well_scorecard  # Assumed to be a dict

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "WELL Scorecard"

    # Set headers
    headers = ["Concept", "Type", "Feature", "Points"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Write data
    for concept, data in well_scorecard.items():
        # Handle Preconditions
        if "Preconditions" in data:
            for item in data["Preconditions"]:
                ws.append([concept, "Precondition", item, "N/A"])
        # Handle Optimizations
        if "Optimizations" in data:
            for feature, points in data["Optimizations"].items():
                ws.append([concept, "Optimization", feature, points])

    # Save to a memory stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Return as downloadable Excel file
    response = HttpResponse(
        file_stream,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=well_scorecard.xlsx'
    return response

def generate_well_certification_insight(project):
    # Gather existing insights from project
    carbon = project.carbon_insight or {}
    waste = project.waste_insight or {}
    design = project.design_insight or {}

    # Convert JSON strings if needed
    import json
    for insight in [carbon, waste, design]:
        if isinstance(insight, str):
            try:
                insight = json.loads(insight)
            except:
                insight = {}

    # Compose the prompt with all insights embedded
    prompt = f"""
    You are an AI assistant helping companies find their WELL v2 Building Standard certification score for WELL v2 certificate tasked with generating a combined insight and forecasting for a construction project.

    You are required to provide proper WELL certification analysis for the following project:{project}

    Please analyze all data and provide a combined JSON output with:

    - A human-readable combined insight summary containing two parts
        1. WELL score for choosen WELL certification by the user and current data available will be value for key "score" and reason for its calculation will be value of key "reasons", reasons value will be list of strings.
        2. Suggestions for improving all factors to maximise WELL score and then predicted WELL score once user follows all suggestion properly. Here also provide result in two keys 'score' and 'reasons'. Similar to above point.
    - In forecasting make sure no past year is provide , forecasting will be made from current year to current year + 1 year
    - In the final insight also add the percentage reduction or increase in cost if adviced material are
     used.
    - Forecasting information for future emissions, waste, and energy efficiency.
    - Graph metrics data suitable for bar graphs, pie charts, or line charts, including:
      - total_carbon_emission (float)
      - predicted_waste_kg (float)
      - energy_efficiency_score (float, 0-100)
      - well_score (int, 0-110) (WELL Certificate score according to your analysis)
      - well_recommendation_grade (str)
    - Scorecard for well will contain all the values provided in the well scorecard json object with its actual values as calculated according to data provided. The total well score calculated should match with that of the scorecard.
    Note: well_recommendation_grade is string providing grades for particular range of score 110 Points [Bronze(0-40 Points),Silver(40-50 Points), Gold(50-60Points),Platinum(60+ Points)],
    Return ONLY the JSON response exactly in this format, with no extra text:

    {{
      "combined_insight": {{
        "current_scenario": {{'score': int, 'reasons': list of strings}},
        "suggestion": {{'score': int, 'reasons': list of strings}}
      }},
      "forecasting": {{
        "carbon_emission_trend": [{{"year": int, "emission": float}}, ...],
        "waste_trend": [{{"year": int, "waste": float}}, ...],
        "energy_efficiency_trend": [{{"year": int, "score": float}}, ...]
      }},
      "graph_metrics": {{
        "total_carbon_emission": float,
        "predicted_waste_kg": float,
        "energy_efficiency_score": float,
        "sustainability_recommendation_score": int,
        "sustainability_recommendation_grade": string
      }},
      "scorecard": {{
            well_scorecard_json
      }}
    }}
    """

    file_path = "LEED Scorecards and Reference Guides/WELL_Full_Scoring_Structure.pdf"

    well_scorecard_json = {
                                "Air": {
                                  "Preconditions": ["A01 - Air Quality", "A02 - Smoke-Free Environment", "A03 - Ventilation Design", "A04 - Construction Pollution Management"],
                                  "Optimizations": {
                                    "A05 - Enhanced Air Quality": 4,
                                    "A06 - Enhanced Ventilation Design": 3,
                                    "A07 - Operable Windows": 2,
                                    "A08 - Air Quality Monitoring and Awareness": 2,
                                    "A09 - Pollution Infiltration Management": 2,
                                    "A10 - Combustion Minimization": 1,
                                    "A11 - Source Separation": 1,
                                    "A12 - Air Filtration": 1,
                                    "A13 - Enhanced Supply Air": 3,
                                    "A14 - Microbe and Mold Control": 1
                                  }
                                },
                                "Water": {
                                  "Preconditions": ["W01 - Water Quality Indicators", "W02 - Drinking Water Quality", "W03 - Basic Water Management"],
                                  "Optimizations": {
                                    "W04 - Enhanced Water Quality": 1,
                                    "W05 - Drinking Water Quality Management": 3,
                                    "W06 - Drinking Water Promotion": 1,
                                    "W07 - Moisture Management": 3,
                                    "W08 - Hygiene Support": 4,
                                    "W09 - Onsite Non-Potable Water Reuse": 2
                                  }
                                },
                                "Nourishment": {
                                  "Preconditions": ["N01 - Fruits and Vegetables", "N02 - Nutritional Transparency"],
                                  "Optimizations": {
                                    "N03 - Refined Ingredients": 2,
                                    "N04 - Food Advertising": 1,
                                    "N05 - Artificial Ingredients": 1,
                                    "N06 - Portion Sizes": 1,
                                    "N07 - Nutrition Education": 1,
                                    "N08 - Mindful Eating": 2,
                                    "N09 - Special Diets": 2,
                                    "N10 - Food Preparation": 1,
                                    "N11 - Responsible Food Sourcing": 1,
                                    "N12 - Food Production": 2,
                                    "N13 - Local Food Environment": 1,
                                    "N14 - Red and Processed Meats": 1
                                  }
                                },
                                "Light": {
                                  "Preconditions": ["L01 - Light Exposure", "L02 - Visual Lighting Design"],
                                  "Optimizations": {
                                    "L03 - Circadian Lighting Design": 3,
                                    "L04 - Electric Light Glare Control": 2,
                                    "L05 - Daylight Design Strategies": 4,
                                    "L06 - Daylight Simulation": 2,
                                    "L07 - Visual Balance": 1,
                                    "L08 - Electric Light Quality": 3,
                                    "L09 - Occupant Lighting Control": 3
                                  }
                                },
                                "Movement": {
                                  "Preconditions": ["V01 - Active Buildings and Communities", "V02 - Ergonomic Workstation Design"],
                                  "Optimizations": {
                                    "V03 - Circulation Network": 3,
                                    "V04 - Facilities for Active Occupants": 3,
                                    "V05 - Site Planning and Selection": 4,
                                    "V06 - Physical Activity Opportunities": 2,
                                    "V07 - Active Furnishings": 2,
                                    "V08 - Physical Activity Spaces and Equipment": 2,
                                    "V09 - Physical Activity Promotion": 1,
                                    "V10 - Self-Monitoring": 1,
                                    "V11 - Ergonomics Programming": 3
                                  }
                                },
                                "Thermal_Comfort": {
                                  "Preconditions": ["T01 - Thermal Performance"],
                                  "Optimizations": {
                                    "T02 - Verified Thermal Comfort": 3,
                                    "T03 - Thermal Zoning": 2,
                                    "T04 - Individual Thermal Control": 3,
                                    "T05 - Radiant Thermal Comfort": 2,
                                    "T06 - Thermal Comfort Monitoring": 1,
                                    "T07 - Humidity Control": 1,
                                    "T08 - Enhanced Operable Windows": 1,
                                    "T09 - Outdoor Thermal Comfort": 3
                                  }
                                },
                                "Sound": {
                                  "Preconditions": ["S01 - Sound Mapping"],
                                  "Optimizations": {
                                    "S02 - Maximum Noise Levels": 3,
                                    "S03 - Sound Barriers": 3,
                                    "S04 - Reverberation Time": 2,
                                    "S05 - Sound Reducing Surfaces": 2,
                                    "S06 - Minimum Background Sound": 2,
                                    "S07 - Impact Noise Management": 3,
                                    "S08 - Enhanced Audio Devices": 2,
                                    "S09 - Hearing Health Conservation": 1
                                  }
                                },
                                "Materials": {
                                  "Preconditions": ["X01 - Material Restrictions", "X02 - Interior Hazardous Materials Management", "X03 - CCA and Lead Management"],
                                  "Optimizations": {
                                    "X04 - Site Remediation": 1,
                                    "X05 - Enhanced Material Restrictions": 2,
                                    "X06 - VOC Restrictions": 4,
                                    "X07 - Materials Transparency": 3,
                                    "X08 - Materials Optimization": 2,
                                    "X09 - Waste Management": 1,
                                    "X10 - Pest Management and Pesticide Use": 1,
                                    "X11 - Cleaning Products and Protocols": 2,
                                    "X12 - Contact Reduction": 2,
                                    "X13 - Fair Labor in Building Products": 3
                                  }
                                },
                                "Mind": {
                                  "Preconditions": ["M01 - Mental Health Promotion", "M02 - Nature and Place"],
                                  "Optimizations": {
                                    "M03 - Mental Health Services": 4,
                                    "M04 - Mental Health Education": 2,
                                    "M05 - Stress Management": 2,
                                    "M06 - Restorative Opportunities": 2,
                                    "M07 - Restorative Spaces": 1,
                                    "M08 - Restorative Programming": 1,
                                    "M09 - Enhanced Access to Nature": 2,
                                    "M10 - Tobacco Cessation": 3,
                                    "M11 - Substance Use Services": 2
                                  }
                                },
                                "Community": {
                                  "Preconditions": ["C01 - Health and Well-Being Promotion", "C02 - Integrative Design", "C03 - Emergency Preparedness", "C04 - Occupant Survey"],
                                  "Optimizations": {
                                    "C05 - Enhanced Occupant Survey": 4,
                                    "C06 - Health Services and Benefits": 5,
                                    "C07 - Enhanced Health and Well-Being Promotion": 2,
                                    "C08 - New Parent Support": 3,
                                    "C09 - New Mother Support": 3,
                                    "C10 - Family Support": 3,
                                    "C11 - Civic Engagement": 2,
                                    "C12 - Talent Recruitment and Workforce Action Plans": 3,
                                    "C13 - Accessibility and Universal Design": 4,
                                    "C14 - Emergency Resources": 2,
                                    "C15 - Emergency Resilience and Recovery": 4,
                                    "C16 - Affordable Housing": 2,
                                    "C17 - Responsible Labor Practices": 3,
                                    "C18 - Support for Victims of Domestic Violence": 2,
                                    "C19 - Education and Support": 2,
                                    "C20 - Historical Acknowledgement": 1,
                                    "C21 - Multisensory Design": 4
                                  }
                                },
                                "Innovation": {
                                  "Optimizations": {
                                    "I01 - Innovate WELL": "Varies",
                                    "I02 - WELL Accredited Professional (WELL AP)": 1,
                                    "I03 - Experience WELL Certification": 1,
                                    "I04 - Gateways to Well-Being": 1,
                                    "I05 - Green Building Rating Systems": 5,
                                    "I06 - Carbon Disclosure and Reduction": 10
                                  }
                                }
                        }
    prompt += f"Here is the example json object for well scorecard json {well_scorecard_json}"
    response = analyze_pdf_from_file(
        file_path=file_path,
        prompt=prompt
    )

    raw_text = response.text.strip()

    # Clean possible markdown or code fences
    if raw_text.startswith("```json"):
        raw_text = raw_text.lstrip("```json").rstrip("```").strip()

    try:
        combined_data = json.loads(raw_text)
    except json.JSONDecodeError as e:
        print("JSON parse error:", e)
        combined_data = {}
    print("insight", combined_data.get("combined_insight"))
    # Save back to project
    project.well_certification_insight = combined_data.get("combined_insight", "")
    project.well_combined_forecasting = combined_data.get("forecasting", {})
    project.well_graph_metrics = combined_data.get("graph_metrics", {})
    project.well_scorecard = combined_data.get("scorecard", {})
    project.save()
    return combined_data


def project_create(request):
    MaterialFormSet = formset_factory(MaterialQuantityForm, extra=1)
    building_types = BuildingType.objects.all()

    if request.method == 'POST':
        formset = MaterialFormSet(request.POST)
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        city = request.POST.get("city")
        country = request.POST.get("country")
        building_type = request.POST.get("building_type")
        well_certification = request.POST.get("well_certification")
        leed_certification = request.POST.get("leed_certification")
        name = request.POST.get("name")
        if formset.is_valid():
            total_emission = 0
            breakdown = []

            for form in formset:
                material = form.cleaned_data.get('material')
                quantity = form.cleaned_data.get('quantity')
                if material and quantity:
                    emission = material.co2_intensity * quantity
                    total_emission += emission
                    breakdown.append({
                        'material': material.name,
                        'quantity': quantity,
                        'emission': round(emission, 2),
                    })

            insight = generate_carbon_footprint_insight(breakdown, total_emission)

            project = Project.objects.create(
                carbon_data=breakdown,
                latitude=float(latitude),
                longitude=float(longitude),
                city=city,
                country=country,
                building_type=building_type,
                carbon_insight=insight,
                name=name,
                leed_certification=leed_certification,
                well_certification=well_certification,
                current_step=2  # Step 1 done, move to step 2
            )
            return redirect('project_step', project_id=project.id, step=1)
    else:
        formset = MaterialFormSet()

    return render(request, 'projects/create.html',
                  {'formset': formset,
                   "building_types": building_types})


# List all projects
def project_list(request):
    projects = Project.objects.all().order_by('-created_at')
    return render(request, 'projects/list.html', {'projects': projects})

# Handle step-wise form filling for project
def project_step(request, project_id, step):
    project = get_object_or_404(Project, id=project_id)
    step = int(step)

    # Prevent access to steps beyond current progress
    if step > project.current_step:
        return redirect('project_step', project_id=project.id, step=project.current_step)

    MaterialFormSet = formset_factory(MaterialQuantityForm, extra=1)
    context = {'project': project, 'step': step}

    if request.method == 'POST':
        # STEP 1: Carbon
        if step == 1:
            formset = MaterialFormSet(request.POST)
            if formset.is_valid():
                total_emission = 0
                breakdown = []

                for form in formset:
                    material = form.cleaned_data.get('material')
                    quantity = form.cleaned_data.get('quantity')
                    if material and quantity:
                        emission = material.co2_intensity * quantity
                        total_emission += emission
                        breakdown.append({
                            'material': material.name,
                            'quantity': quantity,
                            'emission': round(emission, 2),
                        })

                insight = generate_carbon_footprint_insight(breakdown, total_emission)
                project.carbon_data = breakdown
                project.carbon_total_emission = total_emission
                project.carbon_insight = insight

                if project.current_step < 2:
                    project.current_step = 2
                project.save()

                return redirect('project_step', project_id=project.id, step=1)

        # STEP 2: Waste
        elif step == 2:
            form = WasteReductionForm(request.POST)
            if form.is_valid():
                data = form.cleaned_data
                predicted_waste = data['building_size'] * (data['waste_factor'] / 100)

                ai_result = generate_waste_forecasting(
                    data['building_size'], project.carbon_data, data['waste_factor'],
                    data['construction_method'], data['project_phase'], data['climate']
                )

                project.waste_data = {
                    **data,
                    'predicted_waste': predicted_waste
                }
                project.waste_insight = ai_result
                if project.current_step < 3:
                    project.current_step = 3
                project.save()

                return redirect('project_step', project_id=project.id, step=2)

        # STEP 3: Design
        elif step == 3:
            sunlight = float(request.POST.get('sunlight'))
            airflow = float(request.POST.get('airflow'))
            energy_budget = float(request.POST.get('energy_budget'))
            size = project.waste_data.get("building_size")

            suggestion = generate_design_suggestion_gemini(sunlight, airflow, energy_budget, size)

            project.design_data = {
                'sunlight': sunlight,
                'airflow': airflow,
                'energy_budget': energy_budget,
                'size': size
            }
            project.design_insight = suggestion

            if project.current_step < 3:
                project.current_step = 3
            project.save()

            return redirect('project_step', project_id=project.id, step=3)

    else:
        # GET request: render the correct step's form or data
        if step == 1:
            context['formset'] = MaterialFormSet()
            context['materials'] = Material.objects.all()
            context['project'] = project
        elif step == 2:
            context['form'] = WasteReductionForm()
            context['material'] = project.carbon_data
            context['project'] = project
        else:
            context['project'] = project  # contains previous design data if any

    return render(request, 'projects/steps.html', context)

def project_gallery(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    documents = project.documents.all()  # assuming a related name like `documents`
    if request.method == 'POST':
        title = request.POST.get('title')
        file = request.FILES.get('file')

        if title and file:
            Document.objects.create(project=project, title=title, file=file)
            messages.success(request, "Document uploaded successfully!")
            return redirect('project_gallery', project_id=project.id)
        else:
            messages.error(request, "Both title and file are required.")
    return render(request, 'projects/gallery.html', {'project': project, 'documents': documents})


def leed_analysis(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    leed_insight = project.leed_certification_insight
    if not leed_insight:
        generate_leed_certification_insight(project)
    return render(request, 'projects/leed_analysis.html', {'project': project})


def well_analysis(request, project_id):
    project = get_object_or_404(Project, id=project_id)
    well_insight = project.well_certification_insight
    if not well_insight:
        generate_well_certification_insight(project)
    return render(request, 'projects/well_analysis.html', {'project': project})
